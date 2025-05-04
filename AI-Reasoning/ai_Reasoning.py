"""
Program Sistem Rekomendasi Restoran Berbasis Logika Fuzzy
Dibangun untuk memenuhi tugas mata kuliah Sistem Cerdas
oleh:   Devon Arya Daniswara (103022300085)
        Subhan Maulana Ahmad (103022300081)
"""

import openpyxl
import math

# =============================================
# FUNGSI KEANGGOTAAN
# =============================================

def triangular_mf(x, a, b, c):
    """
    Fungsi keanggotaan segitiga
    Parameter:
    x: nilai input
    a: titik awal segitiga
    b: titik puncak segitiga
    c: titik akhir segitiga
    """
    if x <= a or x >= c:
        return 0.0
    elif a < x <= b:
        return (x - a) / (b - a)
    elif b < x < c:
        return (c - x) / (c - b)
    return 0.0

# Fungsi keanggotaan untuk kualitas layanan
def service_poor(x): return triangular_mf(x, 1, 1, 30)
def service_average(x): return triangular_mf(x, 20, 50, 80)
def service_good(x): return triangular_mf(x, 60, 70, 90)
def service_excellent(x): return triangular_mf(x, 80, 100, 100)

# Fungsi keanggotaan untuk harga
def price_cheap(x): return triangular_mf(x, 25000, 25000, 40000)
def price_moderate(x): return triangular_mf(x, 30000, 42500, 55000)
def price_expensive(x): return triangular_mf(x, 45000, 55000, 55000)

# Fungsi keanggotaan untuk output
def output_not_recommended(x): return triangular_mf(x, 0, 0, 30)
def output_maybe_consider(x): return triangular_mf(x, 20, 40, 60)
def output_recommended(x): return triangular_mf(x, 50, 70, 90)
def output_highly_recommended(x): return triangular_mf(x, 80, 100, 100)

# =============================================
# FUZZIFIKASI
# =============================================

def fuzzify_service(service):
    """
    Melakukan fuzzifikasi untuk kualitas layanan
    Mengembalikan dictionary dengan derajat keanggotaan
    """
    return {
        'Buruk': service_poor(service),
        'Sedang': service_average(service),
        'Baik': service_good(service),
        'Sangat Baik': service_excellent(service)
    }

def fuzzify_price(price):
    """
    Melakukan fuzzifikasi untuk harga
    Mengembalikan dictionary dengan derajat keanggotaan
    """
    return {
        'Murah': price_cheap(price),
        'Sedang': price_moderate(price),
        'Mahal': price_expensive(price)
    }

# =============================================
# INFERENSI DAN DEFUZZIFIKASI
# =============================================

def evaluate_rules(service_fuzzy, price_fuzzy):
    """
    Mengevaluasi semua aturan fuzzy dan mengembalikan agregasi output
    """
    # Evaluasi semua aturan menggunakan operator MIN
    rules = [
        min(service_fuzzy['Buruk'], price_fuzzy['Mahal']),    # Aturan 1
        min(service_fuzzy['Buruk'], price_fuzzy['Sedang']),   # Aturan 2
        min(service_fuzzy['Buruk'], price_fuzzy['Murah']),    # Aturan 3
        min(service_fuzzy['Sedang'], price_fuzzy['Mahal']),   # Aturan 4
        min(service_fuzzy['Sedang'], price_fuzzy['Sedang']),  # Aturan 5
        min(service_fuzzy['Sedang'], price_fuzzy['Murah']),   # Aturan 6
        min(service_fuzzy['Baik'], price_fuzzy['Mahal']),     # Aturan 7
        min(service_fuzzy['Baik'], price_fuzzy['Sedang']),    # Aturan 8
        min(service_fuzzy['Baik'], price_fuzzy['Murah']),     # Aturan 9
        min(service_fuzzy['Sangat Baik'], price_fuzzy['Mahal']),  # Aturan 10
        min(service_fuzzy['Sangat Baik'], price_fuzzy['Sedang']), # Aturan 11
        min(service_fuzzy['Sangat Baik'], price_fuzzy['Murah'])   # Aturan 12
    ]
    
    # Kelompokkan aturan berdasarkan output term
    aggregated = {
        'Tidak Direkomendasikan': max(rules[0], rules[1]),
        'Pertimbangkan': max(rules[2], rules[3]),
        'Direkomendasikan': max(rules[4], rules[5], rules[6], rules[9]),
        'Sangat Direkomendasikan': max(rules[7], rules[8], rules[10], rules[11])
    }
    
    return aggregated

def defuzzify(aggregated):
    """
    Melakukan defuzzifikasi menggunakan metode Center of Gravity (COG)
    """
    total_weight = 0.0
    weighted_sum = 0.0
    
    # Sampling pada interval 1-point dari 0 sampai 100
    for x in range(0, 101):
        # Hitung nilai keanggotaan output
        membership = max(
            min(output_not_recommended(x), aggregated['Tidak Direkomendasikan']),
            min(output_maybe_consider(x), aggregated['Pertimbangkan']),
            min(output_recommended(x), aggregated['Direkomendasikan']),
            min(output_highly_recommended(x), aggregated['Sangat Direkomendasikan'])
        )
        
        weighted_sum += x * membership
        total_weight += membership
    
    # Hindari pembagian dengan nol
    return weighted_sum / total_weight if total_weight != 0 else 0

# =============================================
# FUNGSI UTAMA
# =============================================

def process_restaurants(input_file, output_file):
    """
    Fungsi utama untuk memproses data restoran
    """
    try:
        # Baca file input
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        results = []
        
        # Proses setiap baris data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:  # Pastikan ada cukup kolom
                continue
                
            restaurant_id, service, price = row[0], row[1], row[2]
            
            # Fuzzifikasi input
            service_fuzzy = fuzzify_service(service)
            price_fuzzy = fuzzify_price(price)
            
            # Inferensi dan defuzzifikasi
            aggregated = evaluate_rules(service_fuzzy, price_fuzzy)
            score = defuzzify(aggregated)
            
            results.append({
                'id': restaurant_id,
                'service': service,
                'price': price,
                'score': score
            })
        
        # Urutkan berdasarkan skor tertinggi
        results.sort(key=lambda x: x['score'], reverse=True)
        
        # Tampilkan hasil di terminal
        print("\nTOP 5 RESTORAN TERBAIK")
        print("="*60)
        print(f"{'ID':<5}{'Kualitas Layanan':<20}{'Harga (Rp)':<15}{'Skor Rekomendasi':<20}")
        print("-"*60)
        for restaurant in results[:5]:
            print(f"{restaurant['id']:<5}{restaurant['service']:<20}{restaurant['price']:<15}{restaurant['score']:.2f}%")
        print("="*60)
        
        # Simpan ke file output
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.title = "Top 5 Restoran"
        output_sheet.append(['ID', 'Kualitas Layanan', 'Harga (Rp)', 'Skor Rekomendasi (%)'])
        
        for restaurant in results[:5]:
            output_sheet.append([
                restaurant['id'],
                restaurant['service'],
                restaurant['price'],
                round(restaurant['score'], 2)
            ])
        
        output_wb.save(output_file)
        print(f"\nHasil telah disimpan ke file: {output_file}")
        
    except Exception as e:
        print(f"\nError: {str(e)}")

if __name__ == "__main__":
    input_file = "restoran.xlsx"
    output_file = "peringkat.xlsx"
    
    print("SISTEM REKOMENDASI RESTORAN BERBASIS LOGIKA FUZZY")
    print("="*60)
    print(f"Memproses data dari file: {input_file}")
    
    process_restaurants(input_file, output_file)